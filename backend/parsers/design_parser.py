"""Design-document parser — Input #3.

The design document is always an Excel sheet in a fixed 4-column layout:

    S.No | Parameter | Unit | Value

Parsing is fully deterministic (no AI). The only subtlety: Excel silently
coerces some ratio cells typed like "55:45" into time values. Those come back
as datetime.timedelta with a "[h]:mm:ss" number format, so we convert them back
to a "H:MM" ratio string.
"""
from __future__ import annotations

import datetime
from pathlib import Path
from typing import Optional

import openpyxl

from ..schemas import DesignDocExtraction, DesignParam


def _timedelta_to_ratio(td: datetime.timedelta) -> str:
    """[h]:mm formatted time -> ratio string, e.g. 55h45m -> '55:45'."""
    total = int(td.total_seconds())
    h, rem = divmod(total, 3600)
    m, s = divmod(rem, 60)
    return f"{h}:{m:02d}:{s:02d}" if s else f"{h}:{m:02d}"


def _cell_to_str(value, number_format: str) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime.timedelta):
        return _timedelta_to_ratio(value)
    if isinstance(value, datetime.time):
        return f"{value.hour}:{value.minute:02d}"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_design_doc(path: str | Path) -> DesignDocExtraction:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    params: list[DesignParam] = []
    lookup: dict[str, Optional[str]] = {}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        s_no_cell, param_cell, unit_cell, val_cell = (row + (None,) * 4)[:4]

        param = param_cell.value if param_cell else None
        if param is None or str(param).strip() == "":
            continue  # skip blank / separator rows
        param = str(param).strip()

        s_no = s_no_cell.value if s_no_cell else None
        try:
            s_no = int(s_no) if s_no is not None and str(s_no).strip() != "" else None
        except (TypeError, ValueError):
            s_no = None

        unit = unit_cell.value if unit_cell else None
        unit = str(unit).strip() if unit not in (None, "") else None

        value = _cell_to_str(val_cell.value, val_cell.number_format) if val_cell else None

        params.append(DesignParam(s_no=s_no, parameter=param, unit=unit, value=value))
        lookup[param.lower()] = value

    def find(name: str) -> Optional[str]:
        return lookup.get(name.lower())

    return DesignDocExtraction(
        battery_code=find("Battery Code"),
        build_no=find("Build No."),
        design_version=find("Design Version"),
        params=params,
    )
